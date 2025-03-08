import csv
import sys
import os
import re
import openpyxl

def merge_files(file1, file2, output_file):
    # Read the first file (Excel or CSV)
    data1, headers1 = read_file(file1)

    # Read the second file (Excel or CSV)
    data2, headers2 = read_file(file2)

    # Define priority columns (case-insensitive)
    priority_columns = ['ROLL NO.', 'NAME', 'BATCH', 'CLASS', 'ADMISSION DONE BY', 'DOJ']

    # Find the priority column indices (case-insensitive)
    priority_headers = [next((h for h in headers1 if h.upper() == col.upper()), None) for col in priority_columns]
    priority_headers = [h for h in priority_headers if h is not None]

    name_column = next((h for h in priority_headers if h.upper() == 'NAME'), None)

    if not name_column:
        print("Error: NAME column not found in the first file.")
        return

    # Create a dictionary from the second file with Name as the key
    data2_dict = {normalize_name(row[name_column]): row for row in data2}

    # Determine new headers order
    new_headers = (
        priority_headers +
        [h for h in headers2 if h.upper() not in [col.upper() for col in priority_columns] and h not in headers1] +
        [h for h in headers1 if h not in priority_headers]
    )

    # Merge data
    merged_data = []
    for row1 in data1:
        name = normalize_name(row1[name_column])
        merged_row = {header: '' for header in new_headers}
        merged_row.update(row1)
        
        if name in data2_dict:
            # Name found in second file
            row2 = data2_dict[name]
            for key in headers2:
                if key.upper() not in [col.upper() for col in priority_columns] and key not in headers1:
                    merged_row[key] = row2.get(key, '')

        merged_data.append(merged_row)

    # Write the merged data to the output file
    write_file(output_file, new_headers, merged_data)

    # Print debug information
    print(f"Total rows in first file: {len(data1)}")
    print(f"Total rows in second file: {len(data2)}")
    print(f"Total rows in merged data: {len(merged_data)}")
    print(f"New column order: {new_headers}")

def normalize_name(name):
    return re.sub(r'\s+', '', str(name)).lower()

def read_file(file_path):
    _, file_extension = os.path.splitext(file_path)
    if file_extension.lower() == '.xlsx':
        return read_excel(file_path)
    else:
        return read_csv(file_path)

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    headers = [cell.value for cell in sheet[1] if cell.value is not None]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append({k: ('' if v is None else str(v).strip()) for k, v in row_dict.items()})
    return data, headers

def read_csv(file_path):
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        data = [{k: ('' if v is None else str(v).strip()) for k, v in row.items()} for row in reader]
        headers = reader.fieldnames
    return data, headers

def write_file(file_path, headers, data):
    _, file_extension = os.path.splitext(file_path)
    if file_extension.lower() == '.xlsx':
        write_excel(file_path, headers, data)
    else:
        write_csv(file_path, headers, data)

def write_excel(file_path, headers, data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(headers)
    for row in data:
        cell_values = [row.get(header, '') for header in headers]
        sheet.append(cell_values)
    wb.save(file_path)

def write_csv(file_path, headers, data):
    with open(file_path, 'w', newline='', encoding='utf-8') as outfile:
        writer = csv.DictWriter(outfile, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)

def get_file_path(prompt):
    while True:
        file_path = input(prompt).strip()
        # Remove surrounding quotes and ampersand if present
        file_path = re.sub(r'^[&\s"\']+|["\'\s]+$', '', file_path)
        if os.path.isfile(file_path):
            return file_path
        else:
            print(f"Invalid file path: {file_path}")
            print("Please try again.")

if __name__ == "__main__":
    # Get the first file (drag and drop or input)
    if len(sys.argv) > 1:
        file1 = ' '.join(sys.argv[1:])  # Join all arguments in case path contains spaces
        file1 = re.sub(r'^[&\s"\']+|["\'\s]+$', '', file1)  # Remove surrounding quotes and ampersand
    else:
        file1 = get_file_path("Drag and drop the first file or enter its path: ")

    if not os.path.isfile(file1):
        print(f"Invalid file path: {file1}")
        file1 = get_file_path("Please enter a valid file path for the first file: ")

    # Get the second file (input)
    file2 = get_file_path("Enter the path of the second file: ")

    # Get the output file name
    output_file = input("Enter the name for the output file (including extension, e.g., output.xlsx): ")

    merge_files(file1, file2, output_file)
    print(f"Merged data has been written to {output_file}")