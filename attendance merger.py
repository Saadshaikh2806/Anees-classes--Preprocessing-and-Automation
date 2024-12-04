import csv
from collections import defaultdict
import os
import re
import openpyxl
import sys  

# Rest of your code remains the same

def get_file_path(prompt):
    print(prompt)
    print("Please drag and drop the file into this window and press Enter.")
    
    file_path = input().strip()
    
    # Remove quotes if present
    file_path = file_path.strip("\"'")
    
    # Handle Windows-style paths with spaces
    if sys.platform.startswith('win'):
        # Remove any leading "& " that Windows might add
        file_path = re.sub(r'^& ', '', file_path)
        
        # Handle cases where the path is duplicated
        match = re.match(r"([A-Za-z]:\\.*?)'([A-Za-z]:\\.*)", file_path)
        if match:
            file_path = match.group(2)
        
        # Remove any remaining single quotes
        file_path = file_path.replace("'", "")
        
        # If the path starts with a drive letter followed by a colon, it's likely a full path
        if re.match(r'^[a-zA-Z]:', file_path):
            return file_path
        else:
            # If it's not a full path, prepend the current working directory
            return os.path.join(os.getcwd(), file_path)
    else:
        return os.path.abspath(file_path)

def read_attendance_data(file_path):
    attendance_data = {}
    if file_path.lower().endswith('.xlsx'):
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        
        # Get all rows as a list to easily check the header row
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            return attendance_data
            
        # Find the 'Log Records' header row
        log_records_row = None
        emp_code_col = None
        for row_idx, row in enumerate(all_rows):
            for col_idx, cell in enumerate(row):
                if cell and str(cell).strip() == 'Log Records':
                    log_records_row = row_idx
                elif cell and str(cell).strip() == 'Emp Code':
                    emp_code_col = col_idx
            if log_records_row is not None and emp_code_col is not None:
                break
                
        if emp_code_col is None:
            print("Error: 'Emp Code' column not found")
            return attendance_data
            
        # Process each data row
        for row in all_rows[log_records_row + 1:]:  # Skip the header row
            emp_code = row[emp_code_col]
            if emp_code:  # Only process rows with an employee code
                emp_code = str(emp_code).strip()
                
                # Initialize employee record if not exists
                if emp_code not in attendance_data:
                    attendance_data[emp_code] = {'Attendance': {}}
                
                # Collect all time entries from the row
                time_entries = []
                for cell in row:
                    if cell and str(cell).strip():  # Check for non-empty cells
                        cell_value = str(cell).strip()
                        # Skip the employee code and any non-time values
                        if cell_value != emp_code and cell_value != 'Log Records':
                            time_entries.append(cell_value)
                
                # Add all time entries to the employee's record
                for idx, time in enumerate(time_entries, 1):
                    attendance_data[emp_code]['Attendance'][f'Time {idx}'] = time
                
    else:  # Assume CSV if not XLSX
        with open(file_path, 'r') as file:
            csv_reader = csv.reader(file)
            all_rows = list(csv_reader)
            if not all_rows:
                return attendance_data
                
            # Find the 'Log Records' header row
            log_records_row = None
            emp_code_col = None
            for row_idx, row in enumerate(all_rows):
                for col_idx, cell in enumerate(row):
                    if cell and str(cell).strip() == 'Log Records':
                        log_records_row = row_idx
                    elif cell and str(cell).strip() == 'Emp Code':
                        emp_code_col = col_idx
                if log_records_row is not None and emp_code_col is not None:
                    break
                    
            if emp_code_col is None:
                print("Error: 'Emp Code' column not found")
                return attendance_data
                
            # Process each data row
            for row in all_rows[log_records_row + 1:]:  # Skip the header row
                if len(row) > emp_code_col:
                    emp_code = row[emp_code_col]
                    if emp_code:  # Only process rows with an employee code
                        emp_code = str(emp_code).strip()
                        
                        # Initialize employee record if not exists
                        if emp_code not in attendance_data:
                            attendance_data[emp_code] = {'Attendance': {}}
                        
                        # Collect all time entries from the row
                        time_entries = []
                        for cell in row:
                            if cell and str(cell).strip():  # Check for non-empty cells
                                cell_value = str(cell).strip()
                                # Skip the employee code and any non-time values
                                if cell_value != emp_code and cell_value != 'Log Records':
                                    time_entries.append(cell_value)
                        
                        # Add all time entries to the employee's record
                        for idx, time in enumerate(time_entries, 1):
                            attendance_data[emp_code]['Attendance'][f'Time {idx}'] = time
                            
    return attendance_data

def read_all_contact_files():
    contact_info = {}
    data_dir = 'Data'
    for file_name in os.listdir(data_dir):
        file_path = os.path.join(data_dir, file_name)
        if file_name.lower().endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            
            # Get column indices for required fields
            emp_code_idx = next((i for i, h in enumerate(headers) if h.upper() == 'EMP CODE'), None)
            name_idx = next((i for i, h in enumerate(headers) if h.upper() == 'NAME'), None)
            mother_no_idx = next((i for i, h in enumerate(headers) if h.upper() == 'MOTHER NO'), None)
            father_no_idx = next((i for i, h in enumerate(headers) if h.upper() == 'FATHER NO'), None)
            self_no_idx = next((i for i, h in enumerate(headers) if h.upper() == 'SELF NO'), None)
            
            if emp_code_idx is not None and name_idx is not None:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[emp_code_idx]:  # Check if emp code exists
                        emp_code = str(row[emp_code_idx])
                        contact_info[emp_code] = {
                            'NAME': row[name_idx] if name_idx is not None else '',
                            'MOTHER NO': row[mother_no_idx] if mother_no_idx is not None else '',
                            'FATHER NO': row[father_no_idx] if father_no_idx is not None else '',
                            'SELF NO': row[self_no_idx] if self_no_idx is not None else ''
                        }
        elif file_name.lower().endswith('.csv'):
            with open(file_path, 'r') as file:
                csv_reader = csv.DictReader(file)
                fieldnames = [field.strip().upper() for field in csv_reader.fieldnames] if csv_reader.fieldnames else []
                
                if 'EMP CODE' in fieldnames and 'NAME' in fieldnames:
                    for row in csv_reader:
                        emp_code = str(row['EMP CODE']).strip()
                        if emp_code:
                            contact_info[emp_code] = {
                                'NAME': row.get('NAME', '').strip(),
                                'MOTHER NO': row.get('MOTHER NO', '').strip(),
                                'FATHER NO': row.get('FATHER NO', '').strip(),
                                'SELF NO': row.get('SELF NO', '').strip()
                            }
    return contact_info

def merge_data(attendance_data, contact_info):
    merged_data = defaultdict(dict)
    for emp_code, data in attendance_data.items():
        merged_data[emp_code].update(data)
        if emp_code in contact_info:
            merged_data[emp_code].update(contact_info[emp_code])
    return merged_data

def write_merged_data(merged_data, output_file):
    if output_file.lower().endswith('.xlsx'):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Get all unique time columns from the attendance data
        all_times = sorted(set(
            time
            for data in merged_data.values()
            for time in data.get('Attendance', {}).keys()
            if time is not None
        ))

        # Define headers without summary columns
        headers = ['EMP CODE', 'NAME', 'MOTHER NO', 'FATHER NO', 'SELF NO'] + all_times

        sheet.append(headers)

        for emp_code, data in merged_data.items():
            row = [
                emp_code,
                data.get('NAME', ''),
                data.get('MOTHER NO', ''),
                data.get('FATHER NO', ''),
                data.get('SELF NO', '')
            ]
            
            # Add attendance data
            for time in all_times:
                row.append(data.get('Attendance', {}).get(time, ''))
            
            sheet.append(row)

        workbook.save(output_file)
    else:  # Write as CSV
        with open(output_file, 'w', newline='') as file:
            writer = csv.writer(file)
            
            # Get all unique time columns from the attendance data
            all_times = sorted(set(
                time
                for data in merged_data.values()
                for time in data.get('Attendance', {}).keys()
                if time is not None
            ))
            
            # Define headers without summary columns
            headers = ['EMP CODE', 'NAME', 'MOTHER NO', 'FATHER NO', 'SELF NO'] + all_times
            writer.writerow(headers)
            
            for emp_code, data in merged_data.items():
                row = [
                    emp_code,
                    data.get('NAME', ''),
                    data.get('MOTHER NO', ''),
                    data.get('FATHER NO', ''),
                    data.get('SELF NO', '')
                ]
                
                # Add attendance data
                for time in all_times:
                    row.append(data.get('Attendance', {}).get(time, ''))
                
                writer.writerow(row)

    print(f"Merged data has been written to {output_file}")

def main():
    print("Welcome to the Attendance Merger!")
    
    attendance_file = get_file_path("Please drag and drop the attendance file (CSV or XLSX):")
    
    output_file = 'ATTENDANCE_MERGER.xlsx'
    
    print("\nProcessing files...")
    print(f"Attendance file: {attendance_file}")
    
    attendance_data = read_attendance_data(attendance_file)
    contact_info = read_all_contact_files()
    merged_data = merge_data(attendance_data, contact_info)
    write_merged_data(merged_data, output_file)
    
    print(f"\nMerged data has been written to {output_file}")
    print("The merged file is in the same directory as this script.")

if __name__ == "__main__":
    main()
